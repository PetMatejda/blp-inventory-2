import openpyxl

wb = openpyxl.load_workbook('C:/Users/petrm/OneDrive/Dokumenty/BalloonLightGroup/Panalight_2026/BLP_Italy_Rate_Card_v4.xlsx', data_only=True)

with open('C:/Users/petrm/.gemini/antigravity-ide/brain/8bf215bc-e1aa-498e-8057-89a4badbf56a/scratch/rate_card_dump.txt', 'w', encoding='utf-8') as f:
    for name in wb.sheetnames:
        sheet = wb[name]
        f.write(f'========================================\n')
        f.write(f'SHEET: {name} ({sheet.max_row} rows, {sheet.max_column} cols)\n')
        f.write(f'========================================\n')
        for r in range(1, sheet.max_row + 1):
            row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
            if any(row_vals):
                f.write(f'Row {r:2d}: ' + ' | '.join(row_vals) + '\n')
        f.write('\n\n')

print('Dumped all sheets successfully.')
